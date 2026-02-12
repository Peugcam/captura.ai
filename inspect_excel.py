import openpyxl
import sys
import os

# Fix Windows console encoding
if os.name == 'nt':
    import sys
    if sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')

file_path = sys.argv[1] if len(sys.argv) > 1 else "backend/exports/gta_match_20260211_125035_luis.xlsx"

wb = openpyxl.load_workbook(file_path)
print(f"Planilha: {file_path}")
print(f"Abas: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== ABA: {sheet_name} ===")
    print(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas\n")

    # Mostrar primeiras 20 linhas
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 20:
            print(f"Linha {i}: {row}")
        elif i == 21:
            print(f"... ({ws.max_row - 20} linhas restantes)")
            break
    print()
